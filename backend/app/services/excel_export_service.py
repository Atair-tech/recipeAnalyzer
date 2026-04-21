from io import BytesIO
from typing import Iterable, List, Sequence
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def build_excel_bytes(sheet_name: str, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "Sheet1"

    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for row in rows:
        worksheet.append(list(row))

    for column_index, _ in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row in worksheet.iter_rows(min_col=column_index, max_col=column_index):
            cell_value = row[0].value
            if cell_value is None:
                continue
            max_length = max(max_length, len(str(cell_value)))
            row[0].alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def normalize_filename(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value)
    cleaned = cleaned.strip("_")
    return f"{cleaned or fallback}.xlsx"
